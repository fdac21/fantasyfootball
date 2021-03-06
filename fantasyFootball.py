from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook

# create a webdriver and update it
options = Options()
# options.headless = True
driver = webdriver.Chrome(ChromeDriverManager().install(), options=options,)

# determine range of data
years = [year for year in range(2011, 2021)]

# determine positions we want to look at and how many of each
positions = ["QB", "RB", "WR", "TE"]
position_map = {
    "QB": 30,
    "RB": 60,
    "WR": 78,
    "TE": 24
}

# determine an amount of "money" each team has in order to assign value
# 200 dollars given to each of the 12 teams
TOTAL_MONEY = 200 * 12

# store replacement values for each position for each year (10 years * 4 positions)
replacements = [[0, 0, 0, 0] for year in years] 

# store values (initially points then turned into values after some math)
# we want the first 30 QB values each year, first 60 RB values, etc.
values = [[[] for position in positions] for year in years]

# setup excel workbook
wb = Workbook()
ws = wb.active

nextStart = 1

# web scrape replacement value for each position for each of the 10 years
# web scrape points for each player up to replacement player for each of last 10 years
for year in years:
    ws[f"A{nextStart}"] = year
    startRow = nextStart + 2
    currentRow = startRow
    for posIdx, position in enumerate(positions):
        if position == "QB":
            col = "A"
        elif position == "RB":
            col = "B"
        elif position == "WR":
            col = "C"
        else:
            col = "D"
        ws[f"{col}{currentRow}"] = position
        currentRow += 1
        url = f"https://www.footballdb.com/fantasy-football/index.html?pos={position}&yr={year}&wk=all&key=b6406b7aea3872d5bb677f064673c57f"
        driver.get(url)

        points = driver.find_elements_by_xpath("//tbody/tr/td[3]")

        replacement = float(points[position_map[position]].text)
        for i in range(position_map[position]):
            points[i] = float(points[i].text) - replacement
            ws[f"{col}{currentRow}"] = points[i]
            currentRow += 1

        if currentRow > nextStart:
            nextStart = currentRow + 1

        currentRow = startRow

        # add code to get the values up to the replacement value
        # determined by position_map

        # add code to get the replacement value
        # given by position_map[position] + 1
wb.save("values.xlsx")
driver.close()
driver.quit()


# for yearIdx, year in enumerate(years):
#     for posIdx, position in enumerate(positions):
#         # subtracting replacement value from each players points to get value above replacement
#         values[yearIdx][posIdx][:] = [val - replacements[yearIdx][posIdx] for val in values[yearIdx][posIdx]]

# loop through all values above replacement and sum up all values

# divide total by money to get points to dollar ratio

# multiply ratio by each value to get monetary value of each player in an auction draft format

# store this in CSV to show the data we have obtained and make sure the web scraping was a success
