import openpyxl

path = "values.xlsx"

wb = openpyxl.load_workbook(path)
ws = wb.active


positions = ["QB", "RB", "WR", "TE"]
position_map = {
    "QB": 30,
    "RB": 60,
    "WR": 78,
    "TE": 24
}

years = [year for year in range(2011, 2021)]

nextStart = 1

TOTAL_MONEY = 2400

for year in years:
    total = 0
    startRow = nextStart + 2
    for position in positions:
        currentRow = startRow + 1
        if position == "QB":
            col = "A"
        elif position == "RB":
            col = "B"
        elif position == "WR":
            col = "C"
        else:
            col = "D"
        for i in range(position_map[position]):
            total += ws[f"{col}{currentRow}"].value
            currentRow += 1

        if currentRow > nextStart:
            nextStart = currentRow + 1

    dollar_pts_ratio = TOTAL_MONEY / total

    for position in positions:
        if position == "QB":
            col = "A"
            newCol = "F"
        elif position == "RB":
            col = "B"
            newCol = "G"
        elif position == "WR":
            col = "C"
            newCol = "H"
        else:
            col = "D"
            newCol = "I"
        currentRow = startRow
        ws[f"{newCol}{currentRow}"] = position
        currentRow += 1

        for i in range(position_map[position]):
            ws[f"{newCol}{currentRow}"] = round(dollar_pts_ratio * ws[f"{col}{currentRow}"].value, 2)
            currentRow += 1

startRow = nextStart + 1
ws[f"A{startRow}"] = "Avg"
for position in positions:
    valStart = 4
    currentRow = startRow + 1
    if position == "QB":
        col = "A"
        valCol = "F"
    elif position == "RB":
        col = "B"
        valCol = "G"
    elif position == "WR":
        col = "C"
        valCol = "H"
    else:
        col = "D"
        valCol = "I"

    ws[f"{col}{currentRow}"] = position
    currentRow += 1

    for i in range(position_map[position]):
        valRow = valStart + i
        total = 0
        for iteration in range(10):
            total += ws[f"{valCol}{valRow}"].value
            valRow += 82
        ws[f"{col}{currentRow}"] = round(total / 10, 2)
        currentRow += 1


wb.save("values.xlsx")